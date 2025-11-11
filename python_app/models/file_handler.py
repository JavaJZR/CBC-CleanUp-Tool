"""
File Handler Model
Handles file I/O operations and data loading
"""

import pandas as pd
from pathlib import Path
from typing import Optional, Tuple
import re
import openpyxl
from openpyxl import load_workbook


class FileHandler:
    """Handles file I/O operations"""
    
    @staticmethod
    def _read_csv(file_path: str, header: Optional[int] = None) -> pd.DataFrame:
        """Read CSV treating file as trusted and preserving all content."""
        return pd.read_csv(
            file_path,
            header=header,
            dtype=str,
            keep_default_na=False,
            na_filter=False,
            encoding="utf-8-sig"
        )
    
    @staticmethod
    def _read_excel(file_path: str, header: Optional[int] = None) -> pd.DataFrame:
        """Read Excel treating file as trusted and preserving all content."""
        return pd.read_excel(
            file_path,
            header=header,
            dtype=str,
            keep_default_na=False,
            na_filter=False,
            engine="openpyxl"
        )
    
    @staticmethod
    def detect_and_load_csv(file_path: str, file_type: str = None, header_row: Optional[int] = None) -> pd.DataFrame:
        """
        Detect header row and load CSV file ensuring ALL columns are read
        
        Args:
            file_path: Path to the CSV file
            file_type: Type of file ('current_system', 'previous_reference', 'masterlist_current', 'masterlist_resigned')
            header_row: Optional row index (0-based) to use as headers. If None, auto-detect.
        """
        # For masterlists, use the original logic (look for 'Full Name')
        if file_type in ['masterlist_current', 'masterlist_resigned']:
            return FileHandler._load_masterlist_csv(file_path)
        
        # For system reports, use enhanced header detection
        if file_type in ['current_system', 'previous_reference']:
            return FileHandler._load_system_report_csv(file_path, header_row)
        
        # Default behavior for unknown file types
        return FileHandler._load_masterlist_csv(file_path)
    
    @staticmethod
    def _load_masterlist_csv(file_path: str) -> pd.DataFrame:
        """Load CSV file for masterlists (looks for 'Full Name' column)"""
        # First try standard loading with all columns
        try:
            df = FileHandler._read_csv(file_path)
            if 'Full Name' in df.columns:
                return df
        except:
            pass
        
        # Search for 'Full Name' in the first 10 rows
        for header_row in range(10):  # Check first 10 rows
            try:
                df = FileHandler._read_csv(file_path, header=header_row)
                if 'Full Name' in df.columns:
                    return df
            except:
                continue
        
        # If 'Full Name' not found, try keyword detection
        for header_row in range(10):
            try:
                df = FileHandler._read_csv(file_path, header=header_row)
                if FileHandler._is_valid_header(df.columns):
                    return df
            except:
                continue
        
        # Fallback to first row - ensure all columns are read
        return FileHandler._read_csv(file_path)
    
    @staticmethod
    def _load_system_report_csv(file_path: str, header_row: Optional[int] = None) -> pd.DataFrame:
        """Load CSV file for system reports (enhanced header detection)"""
        # If header_row is specified, use it directly
        if header_row is not None:
            try:
                df = FileHandler._read_csv(file_path, header=header_row)
                print(f"Using row {header_row + 1} as headers (user specified)")
                return df
            except Exception as e:
                print(f"Error loading with specified header row {header_row + 1}: {e}, trying automatic detection...")
        
        # Search for system report headers in the first 15 rows
        for header_row in range(15):  # Check first 15 rows for system reports
            try:
                df = FileHandler._read_csv(file_path, header=header_row)
                if FileHandler._is_valid_system_report_header(df.columns):
                    print(f"Found system report headers at row {header_row + 1}")
                    return df
            except:
                continue
        
        # If no valid headers found, try keyword detection
        for header_row in range(15):
            try:
                df = FileHandler._read_csv(file_path, header=header_row)
                if FileHandler._is_valid_header(df.columns):
                    print(f"Found valid headers at row {header_row + 1} using keyword detection")
                    return df
            except:
                continue
        
        # Fallback to first row - ensure all columns are read
        print("Using first row as headers (fallback)")
        return FileHandler._read_csv(file_path)
    
    @staticmethod
    def detect_and_load_excel(file_path: str, file_type: str = None, header_row: Optional[int] = None) -> pd.DataFrame:
        """
        Detect header row and load Excel file ensuring ALL columns are read
        
        Args:
            file_path: Path to the Excel file
            file_type: Type of file ('current_system', 'previous_reference', 'masterlist_current', 'masterlist_resigned')
            header_row: Optional row index (0-based) to use as headers. If None, auto-detect.
        """
        # For masterlists, use the original logic (look for 'Full Name')
        if file_type in ['masterlist_current', 'masterlist_resigned']:
            return FileHandler._load_masterlist_excel(file_path)
        
        # For system reports, use enhanced header detection
        if file_type in ['current_system', 'previous_reference']:
            return FileHandler._load_system_report_excel(file_path, header_row)
        
        # Default behavior for unknown file types
        return FileHandler._load_masterlist_excel(file_path)
    
    @staticmethod
    def _load_masterlist_excel(file_path: str) -> pd.DataFrame:
        """Load Excel file for masterlists (looks for 'Full Name' column)"""
        # First try standard loading with all columns
        try:
            df = FileHandler._read_excel(file_path)
            if 'Full Name' in df.columns:
                return df
        except:
            pass
        
        # Search for 'Full Name' in the first 10 rows
        for header_row in range(10):  # Check first 10 rows
            try:
                df = FileHandler._read_excel(file_path, header=header_row)
                if 'Full Name' in df.columns:
                    return df
            except:
                continue
        
        # If 'Full Name' not found, try keyword detection
        for header_row in range(10):
            try:
                df = FileHandler._read_excel(file_path, header=header_row)
                if FileHandler._is_valid_header(df.columns):
                    return df
            except:
                continue
        
        # If still no valid headers found, try handling merged cells
        print("No valid headers found in masterlist, attempting to handle merged cells...")
        return FileHandler._load_excel_with_merged_cell_handling(file_path)
    
    @staticmethod
    def _load_system_report_excel(file_path: str, header_row: Optional[int] = None) -> pd.DataFrame:
        """Load Excel file for system reports (enhanced header detection)"""
        # If header_row is specified, use it directly
        if header_row is not None:
            try:
                df = FileHandler._read_excel(file_path, header=header_row)
                print(f"Using row {header_row + 1} as headers (user specified)")
                return df
            except Exception as e:
                print(f"Error loading with specified header row {header_row + 1}: {e}, trying automatic detection...")
        
        # Search for system report headers in the first 15 rows
        for header_row in range(15):  # Check first 15 rows for system reports
            try:
                df = FileHandler._read_excel(file_path, header=header_row)
                if FileHandler._is_valid_system_report_header(df.columns):
                    print(f"Found system report headers at row {header_row + 1}")
                    return df
            except:
                continue
        
        # If no valid headers found, try keyword detection
        for header_row in range(15):
            try:
                df = FileHandler._read_excel(file_path, header=header_row)
                if FileHandler._is_valid_header(df.columns):
                    print(f"Found valid headers at row {header_row + 1} using keyword detection")
                    return df
            except:
                continue
        
        # If still no valid headers found, try handling merged cells
        print("No valid headers found, attempting to handle merged cells...")
        return FileHandler._load_excel_with_merged_cell_handling(file_path)
    
    @staticmethod
    def _load_excel_with_merged_cell_handling(file_path: str) -> pd.DataFrame:
        """
        Load Excel file with special handling for merged cells
        This method processes merged cells before pandas reads the file
        """
        try:
            # Load the workbook to handle merged cells
            workbook = load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            # Convert to DataFrame by reading all data and handling merged cells
            data = []
            max_col = worksheet.max_column
            
            for row_num in range(1, worksheet.max_row + 1):
                row_data = []
                for col_num in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    # Handle merged cells by checking if this cell is part of a merged range
                    cell_value = cell.value
                    if cell_value is None:
                        # Check if this cell is part of a merged range
                        for merged_range in worksheet.merged_cells.ranges:
                            if (merged_range.min_row <= row_num <= merged_range.max_row and 
                                merged_range.min_col <= col_num <= merged_range.max_col):
                                # Get the value from the top-left cell of the merged range
                                top_left_cell = worksheet.cell(merged_range.min_row, merged_range.min_col)
                                cell_value = top_left_cell.value
                                break
                    
                    # Convert to string and handle None values
                    row_data.append(str(cell_value) if cell_value is not None else '')
                
                data.append(row_data)
            
            if not data:
                return pd.DataFrame()
            
            # Now try to find valid headers in the processed data
            for header_row in range(min(15, len(data))):
                try:
                    if header_row < len(data):
                        # Create a temporary DataFrame with this row as headers
                        temp_df = pd.DataFrame(data[header_row+1:], columns=data[header_row], dtype=str)
                        
                        # Check for 'Full Name' first (masterlist style)
                        if 'Full Name' in temp_df.columns:
                            print(f"Found 'Full Name' column at row {header_row + 1} after merged cell processing")
                            return temp_df
                        
                        # Check for system report headers
                        if FileHandler._is_valid_system_report_header(temp_df.columns):
                            print(f"Found system report headers at row {header_row + 1} after merged cell processing")
                            return temp_df
                        
                        # Check for general valid headers
                        if FileHandler._is_valid_header(temp_df.columns):
                            print(f"Found valid headers at row {header_row + 1} after merged cell processing")
                            return temp_df
                except:
                    continue
            
            # If no valid headers found, return the DataFrame with first row as headers
            print("Using first row as headers after merged cell processing")
            return pd.DataFrame(data[1:], columns=data[0], dtype=str)
            
        except Exception as e:
            print(f"Error handling merged cells: {e}")
            # Fallback to standard pandas reading
            return FileHandler._read_excel(file_path)
    
    @staticmethod
    def load_file_with_all_columns(file_path: str, file_type: str = None) -> pd.DataFrame:
        """
        Load file ensuring ALL columns are read, especially important for Previous Reference files
        where PERNR might be in Column O or beyond
        
        Args:
            file_path: Path to the file
            file_type: Type of file ('current_system', 'previous_reference', 'masterlist_current', 'masterlist_resigned')
        """
        path_obj = Path(file_path)
        extension = path_obj.suffix.lower()
        
        try:
            if extension == '.csv':
                df = FileHandler._read_csv(file_path)
            elif extension in ['.xlsx', '.xls']:
                df = FileHandler._read_excel(file_path)
            else:
                raise ValueError(f"Unsupported file format: {extension}")
            
            # Log the number of columns loaded for debugging
            print(f"Loaded {len(df.columns)} columns from {path_obj.name}")
            
            return df
            
        except Exception as e:
            print(f"Error loading file {file_path}: {str(e)}")
            # Fallback to original method with file type
            if extension == '.csv':
                return FileHandler.detect_and_load_csv(file_path, file_type)
            else:
                return FileHandler.detect_and_load_excel(file_path, file_type)
    
    @staticmethod
    def _is_valid_system_report_header(columns) -> bool:
        """Check if columns look like valid system report headers"""
        column_names = [str(col).lower() for col in columns]
        
        # Look for system report specific keywords
        system_report_keywords = [
            'pernr', 'pers. number', 'employee number', 'emp number',
            'username', 'user id', 'userid', 'sysid', 'abbreviation',
            'full name', 'name', 'employee name', 'first name', 'last name',
            'user description', 'description', 'user desc', 'desc',
            'department', 'position', 'job title', 'cost center',
            'resignation', 'date', 'effectivity', 'status', 'active',
            'email', 'phone', 'location', 'branch', 'division'
        ]
        
        # Count how many expected keywords are found
        found_keywords = sum(1 for keyword in system_report_keywords 
                           if any(keyword in col for col in column_names))
        
        # For system reports, we need at least 3 keywords to be confident
        # Also check for common system report patterns
        has_user_identifier = any(keyword in ' '.join(column_names) for keyword in ['user', 'username', 'sysid', 'abbreviation'])
        has_employee_data = any(keyword in ' '.join(column_names) for keyword in ['pernr', 'employee', 'name'])
        
        # Additional check: ensure we don't have too many "unnamed" columns
        # which would indicate we're not at the actual header row
        unnamed_columns = sum(1 for col in column_names if 'unnamed' in col)
        total_columns = len(column_names)
        
        # If more than 50% of columns are unnamed, this is likely not a header row
        if unnamed_columns > total_columns * 0.5:
            return False
        
        return found_keywords >= 3 and (has_user_identifier or has_employee_data)
    
    @staticmethod
    def _is_valid_header(columns) -> bool:
        """Check if columns look like valid headers by looking for expected keywords"""
        column_names = [str(col).lower() for col in columns]
        
        # Look for common employee data keywords
        expected_keywords = [
            'pernr', 'pers. number', 'employee number', 'emp number',
            'full name', 'name', 'employee name', 'username',
            'user id', 'userid', 'sysid', 'abbreviation', 
            'user description', 'description', 'user desc', 'desc',
            'department', 'position', 'resignation', 'date', 'effectivity'
        ]
        
        # Count how many expected keywords are found
        found_keywords = sum(1 for keyword in expected_keywords 
                           if any(keyword in col for col in column_names))
        
        # If we find at least 2 expected keywords, consider it a valid header
        return found_keywords >= 2
    
    @staticmethod
    def export_to_excel(df: pd.DataFrame, file_path: str, multi_sheet_data: Optional[dict] = None):
        """Export DataFrame to Excel with optional multi-sheet support, color scheme, and auto-sized columns"""
        if multi_sheet_data:
            # Export with multiple sheets
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, sheet_df in multi_sheet_data.items():
                    if sheet_df is not None and not sheet_df.empty:
                        sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Apply styling and formatting
                        FileHandler._apply_excel_styling(writer, sheet_name, sheet_df)
                        
                        # Apply date formatting to Resignation Date column if it exists
                        if 'Resignation Date' in sheet_df.columns:
                            FileHandler._format_resignation_date_column(writer, sheet_name, sheet_df)
                    else:
                        # Create empty sheet with headers if no data
                        empty_df = pd.DataFrame(columns=df.columns if df is not None else [])
                        empty_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        if not empty_df.empty:
                            FileHandler._apply_excel_styling(writer, sheet_name, empty_df)
        else:
            # Single sheet export
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
                
                # Apply styling and formatting
                FileHandler._apply_excel_styling(writer, 'Sheet1', df)
                
                # Apply date formatting to Resignation Date column if it exists
                if 'Resignation Date' in df.columns:
                    FileHandler._format_resignation_date_column(writer, 'Sheet1', df)
    
    @staticmethod
    def _apply_excel_styling(writer, sheet_name: str, df: pd.DataFrame):
        """Apply color scheme styling and auto-size columns to Excel sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Color scheme
        header_bg_color = "CD1C18"  # Red
        header_text_color = "FFFFFF"  # White
        alternate_row_color = "FFF5F3"  # Light peach
        
        # Style header row (row 1)
        header_fill = PatternFill(start_color=header_bg_color, end_color=header_bg_color, fill_type="solid")
        header_font = Font(bold=True, color=header_text_color, size=11, name="Segoe UI")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Apply alternating row colors and styling
        data_font = Font(size=10, name="Segoe UI")
        data_alignment = Alignment(vertical="top", wrap_text=True)
        
        for row_num in range(2, len(df) + 2):  # Start from row 2 (data rows)
            # Alternate row background color
            if row_num % 2 == 0:
                row_fill = PatternFill(start_color=alternate_row_color, end_color=alternate_row_color, fill_type="solid")
            else:
                row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = row_fill
                cell.font = data_font
                cell.alignment = data_alignment
        
        # Auto-size columns
        for col_num, column_title in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max width
            max_length = len(str(column_title))
            for row_num in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Set column width with some padding (max 50 characters)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = "A2"
    
    @staticmethod
    def _format_resignation_date_column(writer, sheet_name: str, df: pd.DataFrame):
        """Format the Resignation Date column as date type in Excel"""
        from openpyxl.styles import numbers
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Find the column index for "Resignation Date"
        column_index = None
        for col_num, column_title in enumerate(df.columns, 1):
            if column_title == 'Resignation Date':
                column_index = col_num
                break
        
        if column_index is not None:
            # Apply date format to all cells in the Resignation Date column (starting from row 2 to skip header)
            date_format = numbers.FORMAT_DATE_XLSX14  # This is the date format 'mm-dd-yy'
            
            for row_num in range(2, len(df) + 2):  # +2 because Excel is 1-indexed and has header
                cell = worksheet.cell(row=row_num, column=column_index)
                
                # Preserve existing styling (fill and font) when formatting date
                existing_fill = cell.fill
                existing_font = cell.font
                existing_alignment = cell.alignment
                
                # Convert string date to datetime if it's not None/empty
                if cell.value and str(cell.value).strip():
                    try:
                        # Parse the MM/DD/YYYY format
                        from datetime import datetime
                        date_obj = datetime.strptime(str(cell.value), '%m/%d/%Y')
                        cell.value = date_obj
                        cell.number_format = date_format
                        
                        # Restore styling that might have been affected
                        cell.fill = existing_fill
                        cell.font = existing_font
                        cell.alignment = existing_alignment
                    except (ValueError, TypeError):
                        # If date parsing fails, leave as string
                        pass
    
    @staticmethod
    def export_to_csv(df: pd.DataFrame, file_path: str):
        """Export DataFrame to CSV with proper formatting"""
        # Export with proper encoding and formatting
        df.to_csv(file_path, index=False, encoding='utf-8-sig')  # utf-8-sig ensures Excel can open it properly
    
    @staticmethod
    def build_filename(base_name: str, label: str, timestamp: str, extension: str) -> str:
        """Build a standardized filename"""
        # Sanitize base name
        sanitized_base = re.sub(r"[^A-Za-z0-9 _\-]", "", str(base_name)).strip()
        formatted_label = label.replace("_", " ").title()
        return f"{sanitized_base} - {formatted_label} - {timestamp}.{extension}"
    
    @staticmethod
    def get_file_info(file_path: str) -> Tuple[str, str, int, int]:
        """Get file information (name, extension, rows, columns)"""
        path_obj = Path(file_path)
        file_name = path_obj.name
        extension = path_obj.suffix.lower()
        
        # Try to get row and column count
        try:
            if extension == '.csv':
                df = FileHandler._read_csv(file_path)
            elif extension in ['.xlsx', '.xls']:
                df = FileHandler._read_excel(file_path)
            else:
                return file_name, extension, 0, 0
            
            return file_name, extension, len(df), len(df.columns)
        except:
            return file_name, extension, 0, 0
